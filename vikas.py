# importing only those functions
# which are needed
from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinter.ttk import Combobox
import playsound
import gtts
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib



# creating tkinter window
root = Tk()
root.geometry("1486x790")
root.title('Ocean institue of India ')
root.configure(bg='white')
#heading of mu file 
head=Label(
	text="Ocean Institue of India",
	font=('Algerian',40,'bold'),
	bg='white',
	width=80
    
	)
head.pack()
 #sound 

#help menu 
def Help():
	screen=Toplevel(root)
	screen.geometry('350x300+100+100')
	screen.title("Ocean Institue of India")
	screen.resizable(False,False)

	#img help------------------------
	photo = Image.open("D:/A python/wbtech/help.jpg")
	resize__image = photo.resize((300,250))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,image=img,bg='black',border=0)
	label2.photo = img
	label2.pack()

	screen.mainloop()




#home def ______________________
def home():
	user.set("")
	code.set("")





#def section of menu--------------------------------------------------------------------------------------------


#def of apply_ section---------------------------------------------------------
def Apply():
	screen=Toplevel(root)
	screen.title("Ocean Institue of India")
	screen.geometry("1435x750")
	screen.config(bg='white')

	#apply img setting-------------------------
	photo = Image.open("D:/A python/wbtech/zapply1.jpg")
	resize__image = photo.resize((650,650))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,
	image=img,
	bg='black',
	border=0
	)
	label2.photo = img
	label2.place(x=0,y=70)


	#apply frame setting---------------------------
	fr_ame=Frame(screen,width=800,height=683,bg="white")
	fr_ame.place(x=700,y=80)


	#apply heading-------------------------
	head=Label(screen,
	text="Ocean Institue of India Admission Form",
	font=('Algerian',40,'bold'),
	bg='white',
	width=80
	)
	head.pack()


	#apply back menu bar________________________________________
	menu_bar = Menu(screen)
	back=Menu(menu_bar,tearoff=0)
	menu_bar.add_cascade(label="back",menu=back)
	back.add_command(label ='Exit', command = screen.destroy)

	screen.config(menu = menu_bar)


	# clear button
	def clears():
		user_entry.delete(0,'end')
		last_entry.delete(0,'end')
		dateof_entry.delete(0,'end')
		email_entry.delete(0,'end')
		state_entry.delete(0,'end')
		Phone_entry.delete(0,'end')
		Mobile_entry.delete(0,'end')
		passw_entry.delete(0,'end')
		userid_entry.delete(0,'end')
		adhar_entry.delete(0,'end')
		father_entry.delete(0,'end')
		mother_entry.delete(0,'end')



	#excle file_________-----------------------------
	file=pathlib.Path("data.xlsx")
	if file.exists():
		pass
	else:
		file=Workbook()
		sheet=file.active
		sheet["A1"]="Fast Name"
		sheet["B1"]="Last Name"
		sheet["C1"]="Father's Name"
		sheet["D1"]="Mother's Name"
		sheet["E1"]="Gender"
		sheet["F1"]="Date of Birth"
		sheet["G1"]="Course"
		sheet["H1"]="State"
		sheet["I1"]="Adhar"
		sheet["J1"]="Phone"
		sheet["K1"]="Mobile"
		sheet["L1"]="Email"
		sheet["M1"]=" userid"
		sheet["N1"]="password"

		file.save("data.xlsx")





	# apply  mega validation(def)--------------------------------------


	def shave():
		user=user_entry.get()
		last=last_entry.get()
		date=dateof_entry.get()
		gender=gender_com.get()
		email=email_entry.get()
		state=state_entry.get()
		phone=Phone_entry.get()
		mobile=Moblie_entry.get()
		passw=passw_entry.get()
		userid=userid_entry.get()
		course=course_com.get()
		adhar=adhar_entry.get()
		father=father_entry.get()
		mother=mother_entry.get()


		if user=='' or last=='' or date=='' or email=='' or state=='' or phone=='' or mobile=='' or passw=='' or userid=='' or adhar=='' or father=='' or mother=='':
			messagebox.showerror("Invalid ","please input your data")

		elif len(phone) <10 or len(phone) >10 or phone.isalpha()==True:
			messagebox.showerror("Invalid number ","Please type valid number 10 digit")

		elif len(mobile) <10 or len(mobile) >10 or mobile.isalpha()==True:
			messagebox.showerror("Invalid number ","Please type valid number 10 digit")

		elif len(adhar) <12 or len(adhar) >12 or adhar.isalpha()==True:
			messagebox.showerror("Invalid number ","Please type Aadhar number 12 digit")
		elif user.isdigit()==True or last.isdigit()==True or state.isdigit()==True or mother.isdigit()==True or father.isdigit()==True :
			messagebox.showerror("Invalid name ","Please , any name type in letter")
			


		else:
			messagebox.askyesno("sussesful ","Are you sure submit your data")
			file=openpyxl.load_workbook('data.xlsx')
			sheet=file.active
			sheet.cell(column=1,row=sheet.max_row+1,value=user)
			sheet.cell(column=2,row=sheet.max_row,value=last)
			sheet.cell(column=3,row=sheet.max_row,value=father)
			sheet.cell(column=4,row=sheet.max_row,value=mother)
			sheet.cell(column=5,row=sheet.max_row,value=gender)
			sheet.cell(column=6,row=sheet.max_row,value=date)
			sheet.cell(column=7,row=sheet.max_row,value=course)
			sheet.cell(column=8,row=sheet.max_row,value=state)
			sheet.cell(column=9,row=sheet.max_row,value=adhar)
			sheet.cell(column=10,row=sheet.max_row,value=phone)
			sheet.cell(column=11,row=sheet.max_row,value=mobile)
			sheet.cell(column=12,row=sheet.max_row,value=email)
			sheet.cell(column=13,row=sheet.max_row,value=userid)
			sheet.cell(column=14,row=sheet.max_row,value=passw)

			file.save(r'data.xlsx')







		#admission apply frome creat ______----------------------------------------
		#user name first name--------------------
	user_name=Label(fr_ame,text="Fist Name :",font=("times new roman",14,'bold'),bg='white',)
	user_name.grid(row=0,column=1,padx=5,pady=20,sticky=W)

	user_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	user_entry.grid(row=0,column=2,padx=5,pady=20,sticky=W)

	#user name last name--------------------
	last_name=Label(fr_ame,text="Last Name :",font=("times new roman",14,'bold'),bg='white')
	last_name.grid(row=0,column=3,padx=5,pady=20,sticky=W)

	last_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	last_entry.grid(row=0,column=4,padx=5,pady=20,sticky=W)



	#Father name of afather name--------------------
	father_name=Label(fr_ame,text="Father Name :",font=("times new roman",14,'bold'),bg='white')
	father_name.grid(row=1,column=1,padx=5,pady=20,sticky=W)

	father_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	father_entry.grid(row=1,column=2,padx=5,pady=20,sticky=W)

	#mother  name of apply--------------------
	mother_name=Label(fr_ame,text="Mother Name :",font=("times new roman",14,'bold'),bg='white')
	mother_name.grid(row=1,column=3,padx=5,pady=20,sticky=W)

	mother_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	mother_entry.grid(row=1,column=4,padx=5,pady=20,sticky=W)

	# date of birth of user_-------------------------------------------------


	def on_enter(e):
		dateof_entry.delete(0,'end')

	def on_leave(e):
		name=dateof_entry.get()
		if name=='':
			dateof_entry.insert(0,'DD/MM/YYYY')

	dateof_name=Label(fr_ame,text=" Date of Birth :",font=("times new roman",14,'bold'),bg='white')
	dateof_name.grid(row=2,column=1,padx=5,pady=20,sticky=W)

	dateof_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	dateof_entry.grid(row=2,column=2,padx=5,pady=20,sticky=W)
	dateof_entry.insert(0,"DD/MM/YYYY")


	dateof_entry.bind('<FocusIn>',on_enter)
	dateof_entry.bind('<FocusOut>',on_leave)

	#gender name last name--------------------
	gender_name=Label(fr_ame,text="Gender:",font=("times new roman",14,'bold'),bg='white')
	gender_name.grid(row=2,column=3,padx=5,pady=20,sticky=W)



	gender_com=Combobox(fr_ame,values=['Male','Female','Other'],font='arial 14',state='r',width=10)
	gender_com.grid(row=2,column=4,padx=5,pady=20,sticky=W)
	gender_com.set('Male')

	#user state name--------------------
	state_name=Label(fr_ame,text="State Name :",font=("times new roman",14,'bold'),bg='white',)
	state_name.grid(row=3,column=1,padx=5,pady=20,sticky=W)

	state_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	state_entry.grid(row=3,column=2,padx=5,pady=20,sticky=W)

	#user name last name--------------------
	adhar_name=Label(fr_ame,text="Adhar NO. :",font=("times new roman",14,'bold'),bg='white')
	adhar_name.grid(row=3,column=3,padx=5,pady=20,sticky=W)

	adhar_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white',)
	adhar_entry.grid(row=3,column=4,padx=5,pady=20,sticky=W)

	#user user phone number--------------------
	Phone_name=Label(fr_ame,text="Phone No. :",font=("times new roman",14,'bold'),bg='white',)
	Phone_name.grid(row=4,column=1,padx=5,pady=20,sticky=W)

	Phone_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	Phone_entry.grid(row=4,column=2,padx=5,pady=20,sticky=W)

	#user name last name--------------------
	Moblie_name=Label(fr_ame,text="Moblie NO. :",font=("times new roman",14,'bold'),bg='white')
	Moblie_name.grid(row=4,column=3,padx=5,pady=20,sticky=W)

	Moblie_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white',)
	Moblie_entry.grid(row=4,column=4,padx=5,pady=20,sticky=W)


	#user user email id--------------------

	def on_enter(e):
		email_entry.delete(0,'end')

	def on_leave(e):
		name=email_entry.get()
		if name=='':
			email_entry.insert(0,'vikas123@gmail.')


	email_name=Label(fr_ame,text="Email Id :",font=("times new roman",14,'bold'),bg='white',)
	email_name.grid(row=5,column=1,padx=5,pady=10,sticky=W)

	email_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	email_entry.grid(row=5,column=2,padx=5,pady=10,sticky=W)
	email_entry.insert(0,'vikas123@gmail.com')

	email_entry.bind('<FocusIn>',on_enter)
	email_entry.bind('<FocusOut>',on_leave)

	#user name course--------------------
	cours_name=Label(fr_ame,text="Course :",font=("times new roman",14,'bold'),bg='white')
	cours_name.grid(row=5,column=3,padx=5,pady=10,sticky=W)


	course_com=Combobox(fr_ame,values=['Digital marketing','ADCA','MDCA','Video Editer','Graphic Designer'],font='arial 14',state='r',width=15)
	course_com.grid(row=5,column=4,padx=5,pady=20,sticky=W)
	course_com.set('Select your course')

	




	#user userId--------------------
	userid_name=Label(fr_ame,text="UserId :",font=("times new roman",14,'bold'),bg='white',)
	userid_name.grid(row=6,column=1,padx=5,pady=10,sticky=W)

	userid_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white')
	userid_entry.grid(row=6,column=2,padx=5,pady=10,sticky=W)

	#user name course--------------------
	Passw_name=Label(fr_ame,text="Password:",font=("times new roman",14,'bold'),bg='white')
	Passw_name.grid(row=6,column=3,padx=5,pady=10,sticky=W)

	passw_entry=Entry(fr_ame,font=("Arial",13),width=25,bg='white',)
	passw_entry.grid(row=6,column=4,padx=5,pady=10,sticky=W)

	# shave button  apply___------------------------------
	button=Button(fr_ame,width=15,text='Save',bg='blue',fg='white',border=0,font=('Microsoft YaHei UI Light',11,'bold'),command=shave)
	button.grid(row=7,column=2,sticky=W, pady=20)


	button=Button(fr_ame,width=15,text='clear',bg='blue',fg='white',border=0,font=('Microsoft YaHei UI Light',11,'bold'),command=clears)
	button.grid(row=8,column=2,sticky=W)





	screen.mainloop()






	



# result of menu (input your user and password)----------------------------------------------------------------------
def result():
	screen=Toplevel(root)
	screen.title("Ocean Institue of India")
	screen.geometry("300x200+300+300")
	screen.config(bg='black')

	#result menu back menu------------------------
	menu_bar = Menu(screen)
	back=Menu(menu_bar,tearoff=0)
	menu_bar.add_cascade(label="back",menu=back)
	back.add_command(label ='Exit', command = screen.destroy)

	screen.config(menu = menu_bar)



	# resut user----------------------------

	def on_enter(e):
		user.delete(0,'end')
	def on_leave(e):
		name=user.get()
		if name=='':
			user.insert(0,'password')

	user = Entry(screen,width=35,
	fg='black',
	border=2,
	bg='white',
	font=('Microsoft YaHei UI Light',16))
	user.place(x=10,y=10)
	user.insert(0,'username')

	user.bind('<FocusIn>',on_enter)
	user.bind('<FocusOut>',on_leave)

	#password resut-------------------------
	def on_enter(e):
		code.delete(0,'end')
	def on_leave(e):
		name=code.get()
		if name=='':
			code.insert(0,'password')



	code = Entry(screen,width=35,
	fg='black',
	border=2,
	bg='white',
	font=('Microsoft YaHei UI Light',16))
	code.place(x=10,y=65)
	code.insert(0,'password')

	code.bind('<FocusIn>',on_enter)
	code.bind('<FocusOut>',on_leave)


	# result def contion-------------------------
	def vikas():
		username=user.get()
		password=code.get()
		if username=='vikas' and password=='kumar':
			window=Toplevel(root)
			window.title("Ocean Institue of India")
			window.geometry("835x550+300+200")
			window.config(bg="white")
			#result vikas ------------------



			photo = Image.open("D:/A python/wbtech/rvikas.jpg")
			resize__image = photo.resize((500,500))
			img = ImageTk.PhotoImage(resize__image)
			label2 = Label(window,image=img,bg='black',border=0)
			label2.photo = img
			label2.pack()
		elif username=='sunny' and password=='jha':
			window=Toplevel(root)
			window.title("Ocean Institue of India")
			window.geometry("835x550+300+200")
			window.config(bg="white")
			#result sunny ------------------



			photo = Image.open("D:/A python/wbtech/rsunny.jpg")
			resize__image = photo.resize((500,500))
			img = ImageTk.PhotoImage(resize__image)
			label2 = Label(window,image=img,bg='black',border=0)
			label2.photo = img
			label2.pack()
		elif username=='amit' and password=='roy':
			window=Toplevel(root)
			window.title("Ocean Institue of India")
			window.geometry("835x550+300+200")
			window.config(bg="white")
			#result amit ------------------



			photo = Image.open("D:/A python/wbtech/ramit.jpg")
			resize__image = photo.resize((500,500))
			img = ImageTk.PhotoImage(resize__image)
			label2 = Label(window,image=img,bg='black',border=0)
			label2.photo = img
			label2.pack()

			window.mainloop()
		else:
			messagebox.showerror("Invalid ","please input valid password and userid")




			



	#result button-------------------------
	p=Button(screen,width=40,
	pady=7,
	text='sign in',
	bg='blue',
	fg='white',
	border=0,
	command = vikas)
	p.place(x=10,y=110)
	
	screen.mainloop()


#--------------------about call ---------------------------------------

def about():
	screen=Toplevel(root)
	screen.geometry('350x300+100+100')
	screen.title("Ocean Institue of India")
	screen.resizable(False,False)

	photo = Image.open("D:/A python/wbtech/about.jpg")
	resize__image = photo.resize((300,250))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,image=img,bg='black',border=0)
	label2.photo = img
	label2.pack()



	screen.mainloop()


# ---------------call to MDCA fil----------------------

def mdca():
	screen=Toplevel(root)
	screen.geometry('350x300+100+100')
	screen.title("Ocean Institue of India")
	screen.resizable(False,False)

	#img mdca------------------------
	photo = Image.open("D:/A python/wbtech/mdca.jpg")
	resize__image = photo.resize((300,250))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,image=img,bg='black',border=0)
	label2.photo = img
	label2.pack()

	screen.mainloop()


# ---------------call to ADCA fil----------------------

def adca():
	screen=Toplevel(root)
	screen.geometry('350x300+100+100')
	screen.title("Ocean Institue of India")
	screen.resizable(False,False)

	#img adca------------------------
	photo = Image.open("D:/A python/wbtech/adca.jpg")
	resize__image = photo.resize((300,250))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,image=img,bg='black',border=0)
	label2.photo = img
	label2.pack()

	screen.mainloop()


	#---------------them mode dark and light mode______________---------------
def help():
	screen=Toplevel(root)
	screen.geometry('350x300+100+100')
	screen.title("Ocean Institue of India")
	screen.resizable(False,False)

	#img help------------------------
	photo = Image.open("D:/A python/wbtech/help.jpg")
	resize__image = photo.resize((300,250))
	img = ImageTk.PhotoImage(resize__image)
	label2 = Label(screen,image=img,bg='black',border=0)
	label2.photo = img
	label2.pack()

	screen.mainloop()










#image conect root ______________________________________________________________________
image = Image.open("D:/A python/wbtech/zyx1.jpg")

# Resize the image using resize() method----------------------------------------
resize_image = image.resize((900,700 ))

img = ImageTk.PhotoImage(resize_image)

# create label and add resize image
label1 = Label(image=img,
	bg='lightgrey',
	border=0

	)
label1.image = img
label1.place(x=0,y=80)


# user login frame_________________________________________________________________
frame=Frame(root,width=500,height=450,bg="white")
frame.place(x=950,y=100)

#user____________________________________________________________________

def on_enter(e):
	user.delete(0,'end')

def on_leave(e):
	name=user.get()
	if name=='':
		user.insert(0,'username')


user = Entry(frame,width=35,
	fg='black',
	border=0,
	bg='white',
	font=('Microsoft YaHei UI Light',16))
user.place(x=30,y=80)
user.insert(0,'username')

user.bind('<FocusIn>',on_enter)
user.bind('<FocusOut>',on_leave)


Frame(frame,width=300,height=4,bg='black').place(x=30,y=107)

#password ___________________________________________________________________

def on_enter(e):
	code.delete(0,'end')

def on_leave(e):
	name=code.get()
	if name=='':
		code.insert(0,'password')



code = Entry(frame,width=35,
	fg='black',
	border=0,
	bg='white',
	font=('Microsoft YaHei UI Light',16))
code.place(x=30,y=150)
code.insert(0,'password')

code.bind('<FocusIn>',on_enter)
code.bind('<FocusOut>',on_leave)


Frame(frame,width=300,height=4,bg='black').place(x=30,y=177)
# buttun  def condition---------------------------------------------------------------------
def button():
	username=user.get()
	password=code.get()
	if username=='vikas' and password=='kumar':
		screen=Toplevel(root)
		screen.title("Ocean Institue of India")
		screen.geometry("935x550+300+200")
		screen.config(bg="white")
		screen.resizable(False,False)


		Label(screen,text='welcome To Ocean Institue of India',font=('Algerian',36),bg='white').pack()

		#img resume------------------------
		photo = Image.open("D:/A python/wbtech/user.jpg")
		resize__image = photo.resize((450,500))
		img = ImageTk.PhotoImage(resize__image)
		label2 = Label(screen,image=img,bg='white',border=0,)
		label2.photo = img
		label2.place(x=0,y=60)


		#img resume------------------------
		photo = Image.open("D:/A python/wbtech/pass.jpg")
		resize__image = photo.resize((450,500))
		img = ImageTk.PhotoImage(resize__image)
		label2 = Label(screen,image=img,bg='white',border=0,)
		label2.photo = img
		label2.place(x=450,y=60)


		screen.mainloop()

	elif username!='vikas' and password!='kumar':
		messagebox.showerror("Invalid ","Invalid user and pass")

	elif username!='vikas':
		messagebox.showerror("Invalid ","Invalid user")
	elif password!="password":
		messagebox.showerror("Invalid ","Invalid pass")


#button -----------------------------
Button(frame,width=40,
	pady=7,
	text='sign in',
	bg='blue',
	fg='white',
	border=0,
	command=button).place(x=35,y=204)
label=Label(frame,text="Don't have an account?",fg='black',bg="white",font=('Microsoft YaHei UI Light',11,'bold'))
label.place(x=35,y=250)

#sign in __------------------------------------------
sign_up=Button(frame ,width=6,text='sign up',border=0,bg='white',fg='blue',font=('Microsoft YaHei UI Light',11,),command=Apply)
sign_up.place(x=220,y=248)

# Creating Menubar---------------------------------------------------------------
menubar = Menu(root)
#HOME option-------------------------------------------------------------------
home=Menu(menubar,tearoff=0)
menubar.add_cascade(label="Home",menu=home,command=home)
#About option
home=Menu(menubar,tearoff=0)
menubar.add_cascade(label="About",menu=about, command=about)
#Aacademics----------------------------------------------------------------
academic=Menu(menubar,tearoff=0)
menubar.add_cascade(label="Academic",menu=academic)
academic.add_command(label ='MDCA', command = mdca)
academic.add_command(label ='ADCA', command = adca)

academic.add_command(label ='Exit', command = root.destroy)


# Adding File Menu and commands-------------------------------------
file = Menu(menubar, tearoff = 0)
menubar.add_cascade(label ='Admission ', menu = file)
file.add_command(label ='Apply', command = Apply)
file.add_command(label ='Admit Card', command = None)
file.add_command(label ='Result', command = result)
file.add_separator()
file.add_command(label ='Exit', command = root.destroy)

# Adding Edit Menu and commands---------------------------------------------

# Adding Help Menu----------------------------------------------------------------
help_ = Menu(menubar, tearoff = 0)
menubar.add_cascade(label ='Help', menu = help_)
help_.add_command(label =' Help', command =Help)

# display Menu--------------------------------------------------------------------------------
root.config(menu = menubar)




root.mainloop()
